from collections import Counter, defaultdict
from dataclasses import dataclass, replace
from datetime import date, datetime, time
from decimal import Decimal
import os
from pathlib import Path
from time import perf_counter
from typing import Any, Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_HEADERS = [
    "单据编号",
    "流水号",
    "日期",
    "时间",
    "帐簿编号",
    "账簿名称",
    "收款员",
    "收款金额",
    "找零",
    "抹零",
]
RESULT_HEADERS = ["匹配组", "匹配角色", "匹配金额绝对值", *SOURCE_HEADERS]
RULE_KEYWORDS = ("医保", "正收款", "非医保", "负收款", "匹配")
RESULT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 18,
    "C": 16,
    "D": 18,
    "E": 20,
    "F": 12,
    "G": 10,
    "H": 12,
    "I": 14,
    "J": 10,
    "K": 14,
    "L": 10,
    "M": 10,
}


@dataclass
class PaymentMatchResult:
    output_path: Path
    candidate_groups: int
    kept_groups: int
    removed_groups: int
    output_rows: int
    timings: dict[str, float]


@dataclass(frozen=True)
class _SourceRecord:
    values: dict[str, Any]
    cents: int
    occurred_at: datetime
    row_number: int


@dataclass(frozen=True)
class _MatchedRow:
    group_id: int
    role: str
    absolute_cents: int
    record: _SourceRecord


def matches_payment_rule(filename: str | Path, instruction: str) -> bool:
    return Path(filename).suffix.lower() == ".xlsx" and all(
        keyword in instruction for keyword in RULE_KEYWORDS
    )


def _amount_in_cents(value: Any) -> int:
    amount = Decimal(str(value))
    cents = amount * 100
    if cents != cents.to_integral_value():
        raise ValueError("收款金额超过两位小数")
    return int(cents)


def _combined_datetime(date_value: Any, time_value: Any) -> datetime:
    if isinstance(date_value, datetime):
        record_date = date_value.date()
    elif isinstance(date_value, date):
        record_date = date_value
    else:
        record_date = date.fromisoformat(str(date_value))

    if isinstance(time_value, datetime):
        record_time = time_value.time()
    elif isinstance(time_value, time):
        record_time = time_value
    else:
        record_time = time.fromisoformat(str(time_value))
    return datetime.combine(record_date, record_time)


def _find_source_sheet(workbook: Workbook) -> tuple[Worksheet, list[Any]]:
    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        if set(SOURCE_HEADERS).issubset(headers):
            return worksheet, headers
    raise ValueError(f"缺少必要列: {', '.join(SOURCE_HEADERS)}")


def _read_records(worksheet: Worksheet, headers: list[Any]) -> list[_SourceRecord]:
    records = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(value is None for value in values):
            continue
        row = dict(zip(headers, values))
        records.append(
            _SourceRecord(
                values=row,
                cents=_amount_in_cents(row["收款金额"]),
                occurred_at=_combined_datetime(row["日期"], row["时间"]),
                row_number=row_number,
            )
        )
    return records


def _build_matches(records: Iterable[_SourceRecord]) -> tuple[list[_MatchedRow], set[int], int]:
    buckets: dict[tuple[date, int], list[_SourceRecord]] = defaultdict(list)
    for record in records:
        buckets[(record.occurred_at.date(), abs(record.cents))].append(record)

    matched_rows: list[_MatchedRow] = []
    removed_record_ids: set[int] = set()
    candidate_groups = 0
    next_group_id = 1
    for bucket_key in sorted(buckets):
        bucket = sorted(buckets[bucket_key], key=lambda record: record.occurred_at)
        medical_positives = [
            record
            for record in bucket
            if record.values["账簿名称"] == "医保" and record.cents > 0
        ]
        non_medical_negatives = [
            record
            for record in bucket
            if record.values["账簿名称"] != "医保" and record.cents < 0
        ]
        non_medical_positives = [
            record
            for record in bucket
            if record.values["账簿名称"] != "医保" and record.cents > 0
        ]

        for pair_index, (medical, negative) in enumerate(
            zip(medical_positives, non_medical_negatives)
        ):
            candidate_groups += 1
            non_medical_positive = (
                non_medical_positives[pair_index]
                if pair_index < len(non_medical_positives)
                else None
            )
            if non_medical_positive is not None and non_medical_positive.occurred_at < medical.occurred_at:
                removed_record_ids.update((id(medical), id(negative)))
                continue

            matched_rows.extend(
                [
                    _MatchedRow(next_group_id, "医保正收款", abs(medical.cents), medical),
                    _MatchedRow(next_group_id, "非医保负收款", abs(negative.cents), negative),
                ]
            )
            if non_medical_positive is not None:
                matched_rows.append(
                    _MatchedRow(
                        next_group_id,
                        "非医保正收款",
                        abs(non_medical_positive.cents),
                        non_medical_positive,
                    )
                )
            next_group_id += 1
    return matched_rows, removed_record_ids, candidate_groups


def _matched_group_sort_key(
    group_rows: list[_MatchedRow],
) -> tuple[date, datetime, int, int]:
    earliest_row = min(
        group_rows,
        key=lambda row: (row.record.occurred_at, row.record.row_number),
    )
    return (
        earliest_row.record.occurred_at.date(),
        earliest_row.record.occurred_at,
        earliest_row.absolute_cents,
        earliest_row.record.row_number,
    )


def _arrange_matched_rows(matched_rows: list[_MatchedRow]) -> list[_MatchedRow]:
    groups: dict[int, list[_MatchedRow]] = defaultdict(list)
    for row in matched_rows:
        groups[row.group_id].append(row)

    arranged_rows = []
    ordered_groups = sorted(groups.values(), key=_matched_group_sort_key)
    for new_group_id, group_rows in enumerate(ordered_groups, start=1):
        group_rows.sort(
            key=lambda row: (row.record.occurred_at, row.record.row_number)
        )
        arranged_rows.extend(
            replace(row, group_id=new_group_id) for row in group_rows
        )
    return arranged_rows


def _validate_matches(matched_rows: list[_MatchedRow], removed_record_ids: set[int]) -> None:
    if any(id(row.record) in removed_record_ids for row in matched_rows):
        raise ValueError("已剔除匹配组出现在输出中")

    groups: dict[int, list[_MatchedRow]] = defaultdict(list)
    for row in matched_rows:
        groups[row.group_id].append(row)
    for group_rows in groups.values():
        roles = Counter(row.role for row in group_rows)
        if roles["医保正收款"] != 1 or roles["非医保负收款"] != 1:
            raise ValueError("匹配组角色数量不正确")
        if roles["非医保正收款"] > 1:
            raise ValueError("匹配组包含多个非医保正收款")
        if len({row.absolute_cents for row in group_rows}) != 1:
            raise ValueError("匹配组金额不一致")

    expected_group_ids = list(range(1, len(groups) + 1))
    encountered_group_ids = []
    closed_group_ids: set[int] = set()
    current_group_id: int | None = None
    for row in matched_rows:
        if row.group_id == current_group_id:
            continue
        if current_group_id is not None:
            closed_group_ids.add(current_group_id)
        if row.group_id in closed_group_ids:
            raise ValueError("匹配组未连续排列")
        current_group_id = row.group_id
        encountered_group_ids.append(row.group_id)
    if encountered_group_ids != expected_group_ids:
        raise ValueError("匹配组未连续排列")
    for group_rows in groups.values():
        sorted_rows = sorted(
            group_rows,
            key=lambda row: (row.record.occurred_at, row.record.row_number),
        )
        if group_rows != sorted_rows:
            raise ValueError("匹配组内未按日期时间排序")


def _format_result_sheet(result_sheet: Worksheet) -> None:
    dark_teal = "FF164E63"
    white = "FFFFFFFF"
    pale_teal = "FFE8F1F2"
    group_fills = ("FFF4F8F8", "FFFFFFFF")
    divider = Side(style="thin", color="FFD6E1E3")

    for cell in result_sheet[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor=dark_teal)
        cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in result_sheet[2]:
        cell.fill = PatternFill(fill_type="solid", fgColor=pale_teal)
        cell.alignment = Alignment(vertical="center")
    result_sheet["A2"].font = Font(name="Microsoft YaHei", size=10, bold=True, color=dark_teal)
    result_sheet["B2"].font = Font(name="Microsoft YaHei", size=10, color="FF334155")
    result_sheet["B2"].alignment = Alignment(vertical="center", wrap_text=True)

    for cell in result_sheet[3]:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF3F6F7")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    result_sheet["A3"].font = Font(name="Microsoft YaHei", size=10, bold=True, color=dark_teal)

    for cell in result_sheet[4]:
        cell.fill = PatternFill(fill_type="solid", fgColor=dark_teal)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_number in range(5, result_sheet.max_row + 1):
        group_id = result_sheet.cell(row=row_number, column=1).value
        fill_color = group_fills[(int(group_id) - 1) % len(group_fills)]
        for cell in result_sheet[row_number]:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.font = Font(name="Microsoft YaHei", size=10, color="FF1F2937")
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=divider)
        for column in (1, 2, 6, 7, 8, 9, 10):
            result_sheet.cell(row=row_number, column=column).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for column in (3, 11, 12, 13):
            result_sheet.cell(row=row_number, column=column).alignment = Alignment(
                horizontal="right", vertical="center"
            )
        result_sheet.cell(row=row_number, column=6).number_format = "yyyy-mm-dd"
        result_sheet.cell(row=row_number, column=7).number_format = "hh:mm:ss"
        for column in (3, 11, 12, 13):
            result_sheet.cell(row=row_number, column=column).number_format = "0.00"
        result_sheet.row_dimensions[row_number].height = 20

    for column, width in RESULT_COLUMN_WIDTHS.items():
        result_sheet.column_dimensions[column].width = width

    result_sheet.row_dimensions[1].height = 30
    result_sheet.row_dimensions[2].height = 34
    result_sheet.row_dimensions[3].height = 22
    result_sheet.row_dimensions[4].height = 28
    result_sheet.merge_cells("A1:M1")
    result_sheet.merge_cells("B2:M2")
    result_sheet.freeze_panes = "A5"
    result_sheet.auto_filter.ref = f"A4:M{result_sheet.max_row}"
    result_sheet.sheet_view.showGridLines = False


def _write_result_sheet(
    workbook: Workbook,
    matched_rows: list[_MatchedRow],
    candidate_groups: int,
    kept_groups: int,
) -> None:
    if "匹配结果" in workbook.sheetnames:
        del workbook["匹配结果"]
    result_sheet = workbook.create_sheet("匹配结果")
    result_sheet.append(["收款匹配处理结果"])
    result_sheet.append(
        ["规则", "医保正收款和非医保负收款按日期与金额配对；非医保正收款按配对顺序一次性使用。"]
    )
    result_sheet.append(
        [
            "汇总",
            f"候选组 {candidate_groups}",
            f"保留组 {kept_groups}",
            f"剔除组 {candidate_groups - kept_groups}",
            f"输出行 {len(matched_rows)}",
        ]
    )
    result_sheet.append(RESULT_HEADERS)
    for row in matched_rows:
        result_sheet.append(
            [
                row.group_id,
                row.role,
                Decimal(row.absolute_cents) / 100,
                *(row.record.values[header] for header in SOURCE_HEADERS),
            ]
        )
    _format_result_sheet(result_sheet)


def process_payment_workbook(input_path: str | Path, output_dir: str | Path) -> PaymentMatchResult:
    started_at = perf_counter()
    workbook = load_workbook(input_path)
    source_sheet, headers = _find_source_sheet(workbook)
    records = _read_records(source_sheet, headers)
    read_seconds = perf_counter() - started_at

    matching_started_at = perf_counter()
    matched_rows, removed_record_ids, candidate_groups = _build_matches(records)
    matched_rows = _arrange_matched_rows(matched_rows)
    _validate_matches(matched_rows, removed_record_ids)
    match_validate_seconds = perf_counter() - matching_started_at

    export_started_at = perf_counter()
    input_path = Path(input_path)
    output_directory = Path(output_dir)
    output_directory.mkdir(parents=True, exist_ok=True)
    output_path = output_directory / f"{input_path.stem}_匹配结果_{uuid4().hex}.xlsx"
    temporary_path = output_directory / f".{output_path.stem}.{uuid4().hex}.tmp.xlsx"
    kept_groups = len({row.group_id for row in matched_rows})
    _write_result_sheet(workbook, matched_rows, candidate_groups, kept_groups)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    export_seconds = perf_counter() - export_started_at

    return PaymentMatchResult(
        output_path=output_path,
        candidate_groups=candidate_groups,
        kept_groups=kept_groups,
        removed_groups=candidate_groups - kept_groups,
        output_rows=len(matched_rows),
        timings={
            "read_seconds": read_seconds,
            "match_validate_seconds": match_validate_seconds,
            "export_seconds": export_seconds,
            "total_seconds": perf_counter() - started_at,
        },
    )
