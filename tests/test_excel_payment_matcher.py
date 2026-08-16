import os
import tempfile
import unittest
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from customer_service.excel_payment_matcher import (
    _MatchedRow,
    _SourceRecord,
    _validate_matches,
    matches_payment_rule,
    process_payment_workbook,
)


HEADERS = [
    "单据编号",
    "流水号",
    "日期",
    "时间",
    "帐簿编号",
    "账簿名称",
    "收款员",
    "收款金额",
    "找零",
    "抹零",
]


class ExcelPaymentMatcherTests(unittest.TestCase):
    def _record(self, serial, day, clock, ledger_name, amount):
        return [
            f"D{serial}",
            serial,
            day,
            clock,
            "01" if ledger_name == "医保" else "02",
            ledger_name,
            "收款员A",
            amount,
            0,
            0,
        ]

    def _write_source_workbook(self, path, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "收款记录"
        worksheet.append(HEADERS)
        for row in rows:
            worksheet.append(row)
        source_border = Border(bottom=Side(style="thin", color="FF78909C"))
        for cell in worksheet[1]:
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="FF456990")
            cell.border = source_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "@"
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_number, column=3).number_format = "yyyy-mm-dd"
            for column in (8, 9, 10):
                worksheet.cell(row=row_number, column=column).number_format = "0.00"
        worksheet["L1"] = "源页格式"
        worksheet["L1"].font = Font(name="Microsoft YaHei", italic=True, color="FF274C77")
        worksheet["L1"].fill = PatternFill(fill_type="solid", fgColor="FFE7ECEF")
        worksheet["L1"].border = source_border
        worksheet["L1"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet["L1"].number_format = "@"
        worksheet.merge_cells("L1:M1")
        worksheet.row_dimensions[1].height = 27
        worksheet.column_dimensions["A"].width = 23
        worksheet.freeze_panes = "C2"
        worksheet.auto_filter.ref = f"A1:J{worksheet.max_row}"
        worksheet.sheet_view.showGridLines = False

        retained_sheet = workbook.create_sheet("附页")
        retained_sheet["A1"] = "必须保留"
        retained_sheet["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="FF5C374C")
        retained_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFF4E3B2")
        retained_sheet["A1"].border = Border(bottom=Side(style="medium", color="FFB2675E"))
        retained_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        retained_sheet["A1"].number_format = "@"
        retained_sheet.merge_cells("A1:C1")
        retained_sheet.append(["项目", "金额"])
        retained_sheet.append(["样例", 1234.5])
        retained_sheet["B3"].number_format = "0.00"
        retained_sheet.row_dimensions[1].height = 31
        retained_sheet.column_dimensions["A"].width = 25
        retained_sheet.freeze_panes = "A3"
        retained_sheet.auto_filter.ref = "A2:B3"
        retained_sheet.sheet_view.showGridLines = False
        workbook.save(path)

    @staticmethod
    def _color_snapshot(color):
        if color is None:
            return None
        return (
            color.type,
            getattr(color, color.type, None) if color.type else None,
            color.tint,
            color.auto,
        )

    def _cell_style_snapshot(self, cell):
        font = cell.font
        fill = cell.fill
        border = cell.border
        alignment = cell.alignment

        def side_snapshot(side):
            if side is None:
                return None
            return side.style, self._color_snapshot(side.color)

        return (
            (
                font.name,
                font.sz,
                font.bold,
                font.italic,
                font.underline,
                font.strike,
                self._color_snapshot(font.color),
            ),
            (
                fill.fill_type,
                self._color_snapshot(fill.fgColor),
                self._color_snapshot(fill.bgColor),
            ),
            (
                side_snapshot(border.left),
                side_snapshot(border.right),
                side_snapshot(border.top),
                side_snapshot(border.bottom),
                side_snapshot(border.diagonal),
                border.diagonalUp,
                border.diagonalDown,
            ),
            (
                alignment.horizontal,
                alignment.vertical,
                alignment.text_rotation,
                alignment.wrap_text,
                alignment.shrink_to_fit,
                alignment.indent,
            ),
            cell.number_format,
        )

    def _worksheet_snapshot(self, worksheet):
        cells = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cells.append(
                        (
                            cell.coordinate,
                            cell.value,
                            self._cell_style_snapshot(cell),
                        )
                    )
        row_dimensions = tuple(
            sorted(
                (
                    index,
                    dimension.height,
                    dimension.hidden,
                    dimension.outline_level,
                    dimension.collapsed,
                )
                for index, dimension in worksheet.row_dimensions.items()
            )
        )
        column_dimensions = tuple(
            sorted(
                (
                    index,
                    dimension.width,
                    dimension.hidden,
                    dimension.bestFit,
                    dimension.outline_level,
                    dimension.collapsed,
                    dimension.min,
                    dimension.max,
                )
                for index, dimension in worksheet.column_dimensions.items()
            )
        )
        return (
            tuple(cells),
            row_dimensions,
            column_dimensions,
            str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
            worksheet.auto_filter.ref,
            worksheet.sheet_view.showGridLines,
            tuple(sorted(str(cell_range) for cell_range in worksheet.merged_cells.ranges)),
        )

    def _main_rows(self):
        return [
            self._record("M1", date(2026, 7, 1), time(9, 0), "医保", 100),
            self._record("N1", date(2026, 7, 1), time(9, 1), "现金", -100),
            self._record("M2", date(2026, 7, 2), time(10, 0), "医保", 100),
            self._record("N2", date(2026, 7, 2), time(10, 1), "现金", -100),
            self._record("P2", date(2026, 7, 2), time(10, 2), "微信", 100),
            self._record("P3", date(2026, 7, 3), time(11, 0), "微信", 100),
            self._record("M3", date(2026, 7, 3), time(11, 1), "医保", 100),
            self._record("N3", date(2026, 7, 3), time(11, 2), "现金", -100),
        ]

    def test_processes_real_headers_preserves_sheets_and_writes_sorted_matches(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "payments.xlsx"
            output_dir = Path(directory) / "output"
            self._write_source_workbook(source, self._main_rows())

            result = process_payment_workbook(source, output_dir)

            self.assertEqual(result.candidate_groups, 3)
            self.assertEqual(result.kept_groups, 2)
            self.assertEqual(result.removed_groups, 1)
            self.assertEqual(result.output_rows, 5)
            self.assertEqual(
                set(result.timings),
                {"read_seconds", "match_validate_seconds", "export_seconds", "total_seconds"},
            )

            source_workbook = load_workbook(source)
            self.assertNotIn("匹配结果", source_workbook.sheetnames)

            output_workbook = load_workbook(result.output_path, data_only=True)
            self.assertEqual(output_workbook["附页"]["A1"].value, "必须保留")
            result_sheet = output_workbook["匹配结果"]
            self.assertEqual(result_sheet["A1"].value, "收款匹配处理结果")
            self.assertEqual(result_sheet["A2"].value, "规则")
            self.assertEqual(result_sheet["A3"].value, "汇总")
            self.assertEqual(
                [cell.value for cell in result_sheet[4]],
                ["匹配组", "匹配角色", "匹配金额绝对值", *HEADERS],
            )
            output_rows = list(result_sheet.iter_rows(min_row=5, values_only=True))
            self.assertEqual(len(output_rows), 5)
            self.assertEqual(
                Counter(row[1] for row in output_rows),
                {"医保正收款": 2, "非医保负收款": 2, "非医保正收款": 1},
            )
            self.assertEqual({row[2] for row in output_rows}, {100})
            self.assertEqual(
                [(row[5], row[6]) for row in output_rows],
                sorted((row[5], row[6]) for row in output_rows),
            )
            self.assertNotIn("M3", [row[4] for row in output_rows])
            self.assertNotIn("N3", [row[4] for row in output_rows])

    def test_formats_result_sheet_for_readability_without_changing_source_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "formatted.xlsx"
            self._write_source_workbook(source, self._main_rows())
            source_workbook = load_workbook(source)
            original_sheet_snapshots = {
                worksheet.title: self._worksheet_snapshot(worksheet)
                for worksheet in source_workbook.worksheets
            }

            result = process_payment_workbook(source, Path(directory) / "output")

            output_workbook = load_workbook(result.output_path)
            result_sheet = output_workbook["匹配结果"]
            self.assertIn("A1:M1", [str(cell_range) for cell_range in result_sheet.merged_cells.ranges])
            self.assertIn("B2:M2", [str(cell_range) for cell_range in result_sheet.merged_cells.ranges])
            self.assertEqual(result_sheet.freeze_panes, "A5")
            self.assertEqual(result_sheet.auto_filter.ref, "A4:M9")
            self.assertFalse(result_sheet.sheet_view.showGridLines)

            minimum_widths = {
                "A": 18,
                "B": 18,
                "C": 16,
                "D": 18,
                "E": 20,
                "F": 12,
                "G": 10,
                "H": 12,
                "I": 14,
                "J": 10,
                "K": 14,
                "L": 10,
                "M": 10,
            }
            for column, minimum_width in minimum_widths.items():
                self.assertGreaterEqual(result_sheet.column_dimensions[column].width, minimum_width)

            for cell in (result_sheet["A1"], result_sheet["A4"], result_sheet["M4"]):
                self.assertTrue(cell.font.bold)
                self.assertEqual(cell.font.color.rgb, "FFFFFFFF")
                self.assertEqual(cell.fill.fgColor.rgb, "FF164E63")
            self.assertEqual(result_sheet["A1"].alignment.horizontal, "center")
            self.assertTrue(result_sheet["B2"].alignment.wrap_text)

            self.assertEqual(result_sheet["F5"].number_format, "yyyy-mm-dd")
            for coordinate in ("C5", "K5", "L5", "M5"):
                self.assertEqual(result_sheet[coordinate].number_format, "0.00")

            self.assertEqual(result_sheet["A5"].fill.fgColor.rgb, result_sheet["A6"].fill.fgColor.rgb)
            self.assertNotEqual(result_sheet["A5"].fill.fgColor.rgb, result_sheet["A7"].fill.fgColor.rgb)

            self.assertEqual(
                set(output_workbook.sheetnames),
                {*original_sheet_snapshots, "匹配结果"},
            )
            for sheet_name, source_snapshot in original_sheet_snapshots.items():
                with self.subTest(sheet=sheet_name):
                    self.assertEqual(
                        self._worksheet_snapshot(output_workbook[sheet_name]),
                        source_snapshot,
                    )

    def test_outputs_interleaving_pairs_as_contiguous_time_sorted_groups(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "pairs.xlsx"
            self._write_source_workbook(
                source,
                [
                    self._record("M1", date(2026, 7, 4), time(9, 0), "医保", 400),
                    self._record("M2", date(2026, 7, 4), time(9, 1), "医保", 400),
                    self._record("N1", date(2026, 7, 4), time(9, 2), "现金", -400),
                    self._record("N2", date(2026, 7, 4), time(9, 3), "微信", -400),
                ],
            )

            result = process_payment_workbook(source, Path(directory) / "output")

            self.assertEqual((result.candidate_groups, result.kept_groups, result.removed_groups), (2, 2, 0))
            result_sheet = load_workbook(result.output_path, data_only=True)["匹配结果"]
            output_rows = list(result_sheet.iter_rows(min_row=5, values_only=True))
            self.assertEqual([row[0] for row in output_rows], [1, 1, 2, 2])
            self.assertEqual([row[4] for row in output_rows], ["M1", "N1", "M2", "N2"])
            for group_id in (1, 2):
                group_times = [
                    (row[5], row[6]) for row in output_rows if row[0] == group_id
                ]
                self.assertEqual(group_times, sorted(group_times))

    def test_validator_rejects_a_group_that_reappears_after_another_group_starts(self):
        def matched_row(group_id, role, row_number, clock):
            record = _SourceRecord(
                values={"流水号": f"R{row_number}"},
                cents=10000,
                occurred_at=datetime.combine(date(2026, 7, 7), clock),
                row_number=row_number,
            )
            return _MatchedRow(group_id, role, 10000, record)

        interleaved_rows = [
            matched_row(1, "医保正收款", 2, time(9, 0)),
            matched_row(2, "医保正收款", 3, time(9, 1)),
            matched_row(1, "非医保负收款", 4, time(9, 2)),
            matched_row(2, "非医保负收款", 5, time(9, 3)),
        ]

        with self.assertRaisesRegex(ValueError, "匹配组未连续排列"):
            _validate_matches(interleaved_rows, set())

    def test_consumes_each_non_insurance_positive_at_most_once_per_bucket(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "positive-once.xlsx"
            self._write_source_workbook(
                source,
                [
                    self._record("M1", date(2026, 7, 6), time(9, 0), "医保", 500),
                    self._record("M2", date(2026, 7, 6), time(9, 1), "医保", 500),
                    self._record("N1", date(2026, 7, 6), time(9, 2), "现金", -500),
                    self._record("N2", date(2026, 7, 6), time(9, 3), "微信", -500),
                    self._record("P1", date(2026, 7, 6), time(9, 4), "支付宝", 500),
                ],
            )

            result = process_payment_workbook(source, Path(directory) / "output")

            result_sheet = load_workbook(result.output_path, data_only=True)["匹配结果"]
            output_rows = list(result_sheet.iter_rows(min_row=5, values_only=True))
            self.assertEqual(result.output_rows, 5)
            self.assertEqual([row[4] for row in output_rows].count("P1"), 1)

    def test_same_time_non_insurance_positive_does_not_remove_pair(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "same-time.xlsx"
            self._write_source_workbook(
                source,
                [
                    self._record("M1", date(2026, 7, 7), time(9, 0), "医保", 500),
                    self._record("P1", date(2026, 7, 7), time(9, 0), "现金", 500),
                    self._record("N1", date(2026, 7, 7), time(9, 1), "微信", -500),
                ],
            )

            result = process_payment_workbook(source, Path(directory) / "output")

            self.assertEqual((result.candidate_groups, result.kept_groups), (1, 1))
            self.assertEqual(result.output_rows, 3)

    def test_repeated_exports_use_unique_names_and_leave_no_temporary_workbooks(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "payments.xlsx"
            output_dir = Path(directory) / "output"
            self._write_source_workbook(source, self._main_rows())

            first = process_payment_workbook(source, output_dir)
            second = process_payment_workbook(source, output_dir)

            self.assertNotEqual(first.output_path, second.output_path)
            self.assertTrue(first.output_path.is_file())
            self.assertTrue(second.output_path.is_file())
            self.assertRegex(
                first.output_path.name,
                r"^payments_匹配结果_[0-9a-f]{32}\.xlsx$",
            )
            self.assertRegex(
                second.output_path.name,
                r"^payments_匹配结果_[0-9a-f]{32}\.xlsx$",
            )
            self.assertEqual(list(output_dir.glob("*.tmp.xlsx")), [])

    def test_rejects_amounts_with_more_than_two_decimal_places(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "precision.xlsx"
            self._write_source_workbook(
                source,
                [self._record("M1", date(2026, 7, 5), time(9, 0), "医保", "100.001")],
            )

            with self.assertRaisesRegex(ValueError, "超过两位小数"):
                process_payment_workbook(source, Path(directory) / "output")

    def test_matches_payment_rule_requires_xlsx_and_all_rule_keywords(self):
        instruction = "请按医保正收款和非医保负收款完成匹配"

        self.assertTrue(matches_payment_rule("payments.xlsx", instruction))
        self.assertFalse(matches_payment_rule("payments.xls", instruction))
        self.assertFalse(matches_payment_rule("payments.xlsx", "请匹配医保收款"))

    def test_loads_once_and_saves_once(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "spies.xlsx"
            self._write_source_workbook(source, self._main_rows())

            original_load = load_workbook
            original_save = Workbook.save
            with patch(
                "customer_service.excel_payment_matcher.load_workbook", wraps=original_load
            ) as workbook_loader, patch.object(
                Workbook, "save", autospec=True, side_effect=original_save
            ) as workbook_save, patch(
                "customer_service.excel_payment_matcher.os.replace", wraps=os.replace
            ) as atomic_replace:
                result = process_payment_workbook(source, Path(directory) / "output")

            workbook_loader.assert_called_once_with(source)
            self.assertEqual(workbook_save.call_count, 1)
            temporary_path = Path(workbook_save.call_args.args[1])
            self.assertNotEqual(temporary_path, result.output_path)
            self.assertEqual(temporary_path.parent, result.output_path.parent)
            self.assertEqual(temporary_path.suffix, ".xlsx")
            atomic_replace.assert_called_once_with(temporary_path, result.output_path)
            self.assertFalse(temporary_path.exists())


if __name__ == "__main__":
    unittest.main()
